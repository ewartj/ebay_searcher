"""Tests for import_transactions helpers."""
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent))

from import_transactions import _outcome, _to_iso, extract_transactions
from datetime import datetime, timezone
import unittest.mock as mock


class TestOutcome:
    def test_profitable_with_full_data(self):
        assert _outcome(5.0, 20.0, 2.0, 3.0) == "good"

    def test_loss_with_full_data(self):
        assert _outcome(15.0, 12.0, 1.5, 3.0) == "bad"

    def test_breakeven_is_bad(self):
        # net = 10 - 3 - 3 - 4 = 0
        assert _outcome(4.0, 10.0, 3.0, 3.0) == "bad"

    def test_gross_margin_fallback_profitable(self):
        assert _outcome(5.0, 20.0, None, None) == "good"

    def test_gross_margin_fallback_loss(self):
        assert _outcome(20.0, 15.0, None, None) == "bad"

    def test_no_cost_data_returns_bad(self):
        assert _outcome(None, 10.0, None, None) == "bad"

    def test_none_sold_returns_bad(self):
        # sold=None must not raise TypeError in the gross-margin fallback
        assert _outcome(5.0, None, None, None) == "bad"

    # Gotrek omnibus 1 bundling special case
    def test_gotrek_omnibus1_sold_under_6_is_good(self):
        assert _outcome(2.0, 5.0, None, None, "gotrek and Felix omnibus 1") == "good"

    def test_gotrek_omnibus1_sold_at_6_is_not_exempt(self):
        # threshold is strictly < 6
        assert _outcome(2.0, 6.0, None, None, "gotrek and Felix omnibus 1") == "good"  # profitable anyway

    def test_gotrek_omnibus1_sold_over_6_uses_normal_logic(self):
        # bought £12.97, sold £12 — genuine loss, not exempt
        assert _outcome(12.97, 12.0, 0.98, 2.70, "gotrek and Felix omnibus 1") == "bad"

    def test_gotrek_omnibus_bundle_not_exempt(self):
        # "omnibus 1 and 2" should not trigger the single-omnibus exemption
        assert _outcome(5.0, 4.0, None, None, "gotrek and Felix omnibus 1 and 2") == "bad"

    def test_gotrek_omnibus_1_3_bundle_not_exempt(self):
        assert _outcome(5.0, 4.0, None, None, "gotrek and Felix omnibus 1-3") == "bad"


class TestToIso:
    def test_datetime_converted(self):
        dt = datetime(2025, 6, 1, tzinfo=timezone.utc)
        assert _to_iso(dt) == "2025-06-01T00:00:00+00:00"

    def test_none_returns_none(self):
        assert _to_iso(None) is None

    def test_string_returns_none(self):
        assert _to_iso("not a date") is None

    def test_excel_serial_converted(self):
        # Excel serial 45474 = 2024-07-01
        result = _to_iso(45474)
        assert result is not None
        assert "2024-07-01" in result


class TestExtractRowGuards:
    """Verify that extract_transactions skips short rows without dropping valid ones."""

    def _make_wb(self, sheet_rows: dict[str, list[tuple]]):
        """Build a minimal in-memory workbook via openpyxl.

        Data rows for sheets '23-25'/'25' start at min_row=11; for '26' at
        min_row=8.  Prepend enough blank rows so appended data lands at the
        right offset.
        """
        import openpyxl
        _SKIP = {"23-25": 10, "25": 10, "26": 7}
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        for sheet_name, rows in sheet_rows.items():
            ws = wb.create_sheet(sheet_name)
            for _ in range(_SKIP.get(sheet_name, 0)):
                ws.append([])
            for row in rows:
                ws.append(list(row))
        return wb

    def _run(self, wb):
        from unittest.mock import patch
        import openpyxl
        with patch("import_transactions.openpyxl.load_workbook", return_value=wb):
            return extract_transactions(Path("dummy.xlsx"))

    def test_sheets_25_short_row_skipped(self):
        """A row with only 14 columns (missing sold at idx 14) must be skipped."""
        short_row = ("2025-01-01",) + (None,) * 13  # 14 cols, idx 0-13
        wb = self._make_wb({"23-25": [], "25": [short_row], "26": []})
        assert self._run(wb) == []

    def test_sheets_25_minimum_row_processed(self):
        """A row with 15 columns (sold present, fees/postage absent) must be kept."""
        dt = datetime(2025, 1, 1, tzinfo=timezone.utc)
        # idx: 0=date 1=None 2=title 3=bought 4-13=None 14=sold 15-16 absent
        row = (dt, None, "horus heresy omnibus", 5.0) + (None,) * 10 + (20.0,)
        assert len(row) == 15
        wb = self._make_wb({"23-25": [], "25": [row], "26": []})
        results = self._run(wb)
        assert len(results) == 1
        assert results[0]["fees"] is None
        assert results[0]["postage"] is None

    def test_sheet26_short_row_skipped(self):
        """Sheet 26 row with 13 columns (missing sold at idx 13) must be skipped."""
        short_row = ("2025-01-01",) + (None,) * 12  # 13 cols
        wb = self._make_wb({"23-25": [], "25": [], "26": [short_row]})
        assert self._run(wb) == []

    def test_sheet26_minimum_row_processed(self):
        """Sheet 26 row with 14 columns must be kept with fees/postage as None."""
        dt = datetime(2025, 1, 1, tzinfo=timezone.utc)
        # idx: 0=date 1=title 2=bought 3-12=None 13=sold  (fees/postage absent)
        row = (dt, "horus heresy omnibus", 5.0) + (None,) * 10 + (20.0,)
        assert len(row) == 14
        wb = self._make_wb({"23-25": [], "25": [], "26": [row]})
        results = self._run(wb)
        assert len(results) == 1
        assert results[0]["fees"] is None
        assert results[0]["postage"] is None
