"""
One-time import of historical transaction data from ebay.xlsx into the local DB.

Reads sold transactions from sheets '23-25', '25', and '26', extracts
BL/Warhammer-relevant items, and inserts them as:
  - market_prices rows  (sold price = confirmed market value)
  - listing_prices rows (bought price = what was paid)
  - feedback rows       (outcome='good' for profitable sales,
                         outcome='bad' for loss-making sales)

Safe to re-run: aborts if the import has already been done.
Run once: python import_transactions.py [path/to/ebay.xlsx]
"""
import sys
import sqlite3
from datetime import datetime, timezone
from pathlib import Path

import openpyxl

XLSX_PATH = Path(sys.argv[1]) if len(sys.argv) > 1 else Path.home() / "Downloads/ebay.xlsx"

BL_KEYWORDS = [
    "warhammer", "heresy", "horus", "space marine", "eldar",
    "necron", "black library", "gotrek", "felix", "gaunt",
    "eisenhorn", "ravenor", "night lord", "primarch", "siege of terra",
    "ahriman", "word bearer", "iron warrior", "death guard",
    "void stalker", "soul hunter", "blood reaver", "ultramar", "ultramarine",
    "dark angel", "blood angel", "space wolf", "grey knight",
    "inquisitor", "mechanicum", "mechanicus", "adeptus",
    "cadia", "sabbat", "tanith",
    "malus darkblade", "genevieve", "sigmar", "archaon", "nagash",
    "black legion", "talon of horus", "clonelord", "manflayer",
    "lords of silence", "broken city", "deathwatch", "macharian",
    "nightbringer", "titanicus", "double eagle", "helsreach",
    "storm of iron", "brothers of the snake", "daemon world",
    "soul drinker", "blood raven", "enforcer", "salamander",
    "tome of fire", "bastion wars", "cassius", "word bearers",
    "the red path", "soul wars", "vampire genevieve", "drakenfels",
    "thunder and steel", "elves omnibus", "knights of caliban",
    "legacy of caliban", "lord of the night", "wrath of iron",
    "the magos", "praetorian of dorn", "master of sanctity",
    "beast arises", "witch finder",
]


def is_bl(title: str) -> bool:
    t = title.lower()
    return any(kw in t for kw in BL_KEYWORDS)


def _outcome(bought, sold, fees, postage, title: str = "") -> str:
    """Return 'good' or 'bad' based on whether the transaction was profitable."""
    # Gotrek omnibus 1 sold cheaply (<£6) is intentional — kept for bundling
    # with omnibuses 2 & 3 where the bundle fetches £38-82.
    t = title.lower()
    if (
        "gotrek" in t
        and "omnibus 1" in t
        and "omnibus 1 and" not in t
        and "omnibus 1-" not in t
        and sold < 6
    ):
        return "good"

    if not isinstance(sold, (int, float)):
        return "bad"

    # If we have full cost data, use net profit
    if all(isinstance(v, (int, float)) for v in [bought, fees, postage]):
        net = sold - fees - postage - bought
        return "good" if net > 0 else "bad"
    # Fall back to gross margin when fees/postage unknown
    if isinstance(bought, (int, float)):
        return "good" if sold > bought else "bad"
    # No cost data at all — can't determine profitability
    return "bad"


def extract_transactions(path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(str(path))
    rows = []

    # Sheets 23-25 and 25:
    # title=C(2), bought=D(3), sold=O(14), fees=P(15), postage=Q(16),
    # bought_date=A(0), sold_date=L(11)
    for sheet_name in ("23-25", "25"):
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=11, max_row=ws.max_row, values_only=True):
            if len(row) < 15:  # need at least up to sold (idx 14)
                continue
            title = row[2]
            bought      = row[3]
            sold        = row[14]
            fees        = row[15] if len(row) > 15 else None
            postage     = row[16] if len(row) > 16 else None
            bought_date = row[0]
            sold_date   = row[11]
            if not (title and isinstance(title, str)):
                continue
            if sold and isinstance(sold, (int, float)) and sold > 0:
                rows.append({
                    "title":      title.strip(),
                    "bought":     bought  if isinstance(bought,  (int, float)) else None,
                    "sold":       float(sold),
                    "fees":       fees    if isinstance(fees,    (int, float)) else None,
                    "postage":    postage if isinstance(postage, (int, float)) else None,
                    "bought_date": bought_date,
                    "sold_date":   sold_date,
                })

    # Sheet 26:
    # title=B(1), bought=C(2), sold=N(13), fees=O(14), postage=P(15),
    # bought_date=A(0), sold_date=K(10)
    ws26 = wb["26"]
    for row in ws26.iter_rows(min_row=8, max_row=ws26.max_row, values_only=True):
        if len(row) < 14:  # need at least up to sold (idx 13)
            continue
        title = row[1]
        bought      = row[2]
        sold        = row[13]
        fees        = row[14] if len(row) > 14 else None
        postage     = row[15] if len(row) > 15 else None
        bought_date = row[0]
        sold_date   = row[10]
        if not (title and isinstance(title, str)):
            continue
        if sold and isinstance(sold, (int, float)) and sold > 0:
            rows.append({
                "title":      title.strip(),
                "bought":     bought  if isinstance(bought,  (int, float)) else None,
                "sold":       float(sold),
                "fees":       fees    if isinstance(fees,    (int, float)) else None,
                "postage":    postage if isinstance(postage, (int, float)) else None,
                "bought_date": bought_date,
                "sold_date":   sold_date,
            })

    return [r for r in rows if is_bl(r["title"])]


def _to_iso(dt) -> str | None:
    if isinstance(dt, datetime):
        return dt.replace(tzinfo=timezone.utc).isoformat()
    # Excel sometimes stores dates as integer serials (days since 1899-12-30)
    if isinstance(dt, (int, float)):
        try:
            from datetime import timedelta
            return (datetime(1899, 12, 30, tzinfo=timezone.utc) + timedelta(days=dt)).isoformat()
        except (ValueError, OverflowError):
            pass
    return None


def already_imported(db_path: str) -> bool:
    conn = sqlite3.connect(db_path)
    row = conn.execute(
        "SELECT value FROM settings WHERE key = 'transactions_imported'"
    ).fetchone()
    conn.close()
    return row is not None and row[0] == "1"


def seed_db(transactions: list[dict], db_path: str) -> None:
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row

    cur = conn.execute(
        "INSERT INTO scans (scanned_at) VALUES (?)",
        (datetime(2024, 1, 1, tzinfo=timezone.utc).isoformat(),),
    )
    scan_id = cur.lastrowid

    market_rows = []
    listing_rows = []
    feedback_rows = []
    seen_feedback: set[str] = set()

    for tx in transactions:
        sold_at   = _to_iso(tx["sold_date"])
        bought_at = _to_iso(tx["bought_date"])
        title = tx["title"]

        # Only insert into date-range-queried tables when the date is known;
        # an epoch sentinel would silently corrupt trend/history queries.
        if sold_at is not None:
            market_rows.append((scan_id, title, tx["sold"], "transaction_history", sold_at))

        if tx["bought"] is not None and bought_at is not None:
            listing_rows.append((scan_id, title, tx["bought"], "transaction_history", bought_at))

        key = title.lower()
        if key not in seen_feedback:
            seen_feedback.add(key)
            url = f"tx://{key.replace(' ', '-')}"
            outcome = _outcome(tx["bought"], tx["sold"], tx["fees"], tx["postage"], title)
            feedback_rows.append((url, title, outcome, sold_at or "unknown"))

    conn.executemany(
        "INSERT INTO market_prices (scan_id, title, market_price, price_source, scanned_at) "
        "VALUES (?, ?, ?, ?, ?)",
        market_rows,
    )
    conn.executemany(
        "INSERT INTO listing_prices (scan_id, title, price_gbp, source, scanned_at) "
        "VALUES (?, ?, ?, ?, ?)",
        listing_rows,
    )
    conn.executemany(
        "INSERT INTO feedback (url, title, outcome, created_at) VALUES (?, ?, ?, ?)",
        feedback_rows,
    )
    conn.execute(
        "INSERT OR REPLACE INTO settings (key, value) VALUES ('transactions_imported', '1')"
    )
    conn.commit()
    conn.close()

    good = sum(1 for _, _, o, _ in feedback_rows if o == "good")
    bad  = sum(1 for _, _, o, _ in feedback_rows if o == "bad")
    print(f"Inserted {len(market_rows)} market_prices rows")
    print(f"Inserted {len(listing_rows)} listing_prices rows")
    print(f"Inserted {good} good + {bad} bad feedback titles")


if __name__ == "__main__":
    sys.path.insert(0, str(Path(__file__).parent))
    import config  # noqa — loads DB_PATH from env/.env
    import db as db_module

    db_module.init_db()

    if already_imported(config.DB_PATH):
        print("Already imported — skipping. Delete the 'transactions_imported' "
              "settings key to re-run.")
        sys.exit(0)

    print(f"Reading: {XLSX_PATH}")
    txns = extract_transactions(XLSX_PATH)
    print(f"Found {len(txns)} BL/Warhammer transactions")
    for t in txns:
        print(f"  {t['title'][:55]:<55} bought=£{t['bought'] or '?':<6} sold=£{t['sold']:.2f}")

    print(f"\nSeeding DB: {config.DB_PATH}")
    seed_db(txns, config.DB_PATH)
    print("Done.")
